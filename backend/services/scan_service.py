"""File scan engine - recursive scan, SHA256 dedup, category stats, change detection.

Runnable as standalone script: python scan_service.py --library-id 1 --path /target/dir
Outputs progress JSON lines to stdout, final result as last line.

Also usable as a JSON-RPC service module via execute(method, params).
"""

import sys
import os
import json
import hashlib
import time
import argparse
import re
from pathlib import Path

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from backend.db.connection import get_connection

EXCEL_EXT = {'.xlsx', '.xls', '.xlsm', '.xlsb'}
IMAGE_EXT = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg', '.ico', '.heic', '.heif'}
CAD_EXT = {'.dwg', '.dxf', '.step', '.stp', '.iges', '.igs', '.prt', '.asm',
           '.sldprt', '.sldasm', '.catpart', '.catproduct', '.ipt', '.iam', '.stl', '.obj'}
PDF_EXT = {'.pdf'}

_OCR_ATTEMPTED = False
_OCR_AVAILABLE = False


def _try_ocr_image(image_path: str) -> str:
    """Try to OCR an image file. Returns recognized text or empty string on any failure."""
    global _OCR_ATTEMPTED, _OCR_AVAILABLE
    if _OCR_ATTEMPTED and not _OCR_AVAILABLE:
        return ""
    _OCR_ATTEMPTED = True
    try:
        from backend.services.ocr_service import _load_ocr, _recognize_text
        _load_ocr()
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        results = _recognize_text(image_bytes)
        _OCR_AVAILABLE = True
        return " ".join(r["text"] for r in results)
    except Exception:
        _OCR_AVAILABLE = False
        return ""


def _add_activity_log(level: str, source: str, message: str):
    try:
        conn = get_connection()
        conn.execute(
            "INSERT INTO activity_logs (level, source, message) VALUES (?, ?, ?)",
            (level, source, message),
        )
        conn.commit()
    except Exception:
        pass


def classify_file(ext: str) -> str:
    e = ext.lower()
    if e in EXCEL_EXT: return "excel"
    if e in IMAGE_EXT: return "image"
    if e in CAD_EXT: return "cad"
    if e in PDF_EXT: return "pdf"
    return "other"


def sha256_file(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


_scan_start_time: float = 0.0


def emit_progress(phase: str, current: int, total: int, current_file: str = ""):
    pct = int(current / max(total, 1) * 100)
    elapsed = round(time.time() - _scan_start_time, 1) if _scan_start_time > 0 else 0.0
    if pct > 0 and elapsed > 0:
        eta = round(elapsed / pct * (100 - pct), 1)
    else:
        eta = 0.0
    payload = {
        "type": "progress",
        "phase": phase,
        "current": current,
        "total": total,
        "current_file": current_file,
        "percent": pct,
        "elapsed_sec": elapsed,
        "eta_sec": eta,
    }
    sys.stdout.write(json.dumps(payload, ensure_ascii=False) + "\n")
    sys.stdout.flush()


def emit_result(result: dict):
    result["type"] = "result"
    sys.stdout.write(json.dumps(result, ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _insert_image_record(conn, filepath: str, info: dict):
    folder = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))

    conn.execute(
        """INSERT OR REPLACE INTO images
           (img_id, source_type, file_path, folder, filename, size_bytes, width, height, file_hash, status, last_modified, indexed_at)
           VALUES (?, 'file_image', ?, ?, ?, ?, NULL, NULL, ?, 'normal', ?, ?)""",
        (info["hash"], filepath, folder, filename, info["size"], info["hash"], mtime, now)
    )


def _insert_cad_record(conn, filepath: str, info: dict):
    folder = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    ext = info["ext"].lower().lstrip(".")
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))
    cad_id = f"CAD-{info['hash'][:16]}"

    conn.execute(
        """INSERT OR REPLACE INTO cad_files
           (cad_id, file_path, folder, filename, extension, size_bytes, file_hash, status, last_modified, indexed_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'normal', ?, ?)""",
        (cad_id, filepath, folder, filename, ext, info["size"], info["hash"], mtime, now)
    )


def _count_pdf_pages(filepath: str) -> int:
    """Quick PDF page count by searching for /Type/Page objects."""
    try:
        import re
        with open(filepath, 'rb') as f:
            content = f.read()
        pages = len(re.findall(br'/Type\s*/Page[^s]', content))
        return pages if pages > 0 else 0
    except Exception:
        return 0


def _insert_pdf_record(conn, filepath: str, info: dict):
    folder = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))
    doc_id = f"DOC-{info['hash'][:16]}"
    page_count = _count_pdf_pages(filepath)

    conn.execute(
        """INSERT OR REPLACE INTO pdf_files
           (doc_id, file_path, folder, filename, size_bytes, page_count, file_hash, status, last_modified, indexed_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'normal', ?, ?)""",
        (doc_id, filepath, folder, filename, info["size"], page_count, info["hash"], mtime, now)
    )


def _extract_excel_images(filepath: str, library_path: str) -> int:
    """Extract embedded images from Excel. Returns count of images extracted."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write(f"[scan] openpyxl not installed, skipping Excel images: {filepath}\n")
        return 0

    extracted_dir = os.path.join(_PROJECT_ROOT, "backend", "data", "extracted_images")
    os.makedirs(extracted_dir, exist_ok=True)

    conn = get_connection()
    folder = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    wb = load_workbook(filepath, data_only=True)
    count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        images = getattr(ws, '_images', [])
        if not images:
            continue

        for img in images:
            try:
                # openpyxl >= 3.1 dropped .image; use _data() as fallback
                if hasattr(img, 'image'):
                    pil_img = img.image
                else:
                    from io import BytesIO
                    from PIL import Image as _PILImage
                    pil_img = _PILImage.open(BytesIO(img._data()))
                if pil_img is None:
                    continue

                width = pil_img.width
                height = pil_img.height

                # Determine extension from PIL image format
                fmt = pil_img.format if pil_img.format else "PNG"
                ext = fmt.lower()
                if ext == "jpeg":
                    ext = "jpg"

                # Generate sequential IMG ID
                max_row = conn.execute(
                    "SELECT img_id FROM images WHERE img_id LIKE 'IMG-%' ORDER BY img_id DESC LIMIT 1"
                ).fetchone()
                if max_row:
                    last_num = int(max_row["img_id"].split("-")[1])
                    img_id = f"IMG-{last_num + 1:06d}"
                else:
                    img_id = "IMG-000001"

                # Save image file
                img_filename = f"{img_id}.{ext}"
                img_path = os.path.join(extracted_dir, img_filename)
                pil_img.save(img_path, format=fmt)

                img_size = os.path.getsize(img_path)

                # Try to determine row from anchor for proper EX-ID linking
                anchor_row = None
                if hasattr(img, 'anchor'):
                    anchor = img.anchor
                    # openpyxl anchor types: OneCellAnchor, TwoCellAnchor, AbsoluteAnchor
                    if hasattr(anchor, '_from'):
                        anchor_row = anchor._from.row + 1  # 0-indexed → 1-indexed
                    elif hasattr(anchor, 'row'):
                        anchor_row = anchor.row + 1

                # Find matching excel_records entry
                ex_ref = ''
                if anchor_row is not None:
                    ex_row = conn.execute(
                        "SELECT ex_id FROM excel_records WHERE file_path = ? AND sheet_name = ? AND row_number = ? LIMIT 1",
                        (filepath, sheet_name, anchor_row)
                    ).fetchone()
                    if ex_row:
                        ex_ref = ex_row["ex_id"]

                # Fallback: link to first record for this file+sheet
                if not ex_ref:
                    ex_row = conn.execute(
                        "SELECT ex_id FROM excel_records WHERE file_path = ? AND sheet_name = ? LIMIT 1",
                        (filepath, sheet_name)
                    ).fetchone()
                    if ex_row:
                        ex_ref = ex_row["ex_id"]

                ocr_text = _try_ocr_image(img_path)
                conn.execute(
                    """INSERT OR REPLACE INTO images
                       (img_id, source_type, file_path, folder, filename, size_bytes, width, height, file_hash, ex_ref, ocr_text, status, last_modified, indexed_at)
                       VALUES (?, 'excel_embedded', ?, ?, ?, ?, ?, ?, ?, ?, ?, 'normal', ?, ?)""",
                    (img_id, filepath, folder, img_filename, img_size, width, height, None, ex_ref, ocr_text, mtime, now)
                )

                count += 1
            except Exception as e:
                sys.stderr.write(f"[scan] excel image extract error: {filepath} sheet={sheet_name}: {e}\n")
                sys.stderr.flush()
                continue

    wb.close()
    return count


def _log_change(conn, change_type: str, filepath: str, old_value: str = None, new_value: str = None):
    conn.execute(
        """INSERT INTO change_logs (change_type, file_path, old_value, new_value, status, created_at)
           VALUES (?, ?, ?, ?, 'processed', datetime('now','localtime'))""",
        (change_type, filepath, old_value, new_value)
    )


def _auto_associate(conn, library_path: str):
    """Auto-associate images with CAD files, Excel records, and PDFs.

    Creates match records with status 'auto' for:
    - Images and CAD files in the same folder
    - Images and CAD files with matching base filenames
    - Excel embedded images with their Excel records
    - PDFs with similar-named images
    """
    matches_created = 0

    # ── Same-folder: images ↔ CAD ──
    rows = conn.execute(
        """SELECT i.img_id, i.folder, i.filename, c.cad_id, c.filename as cad_filename
           FROM images i
           JOIN cad_files c ON i.folder = c.folder
           WHERE i.source_type = 'file_image'
             AND i.img_id NOT IN (SELECT img_id FROM matches WHERE cad_id = c.cad_id)
           LIMIT 500"""
    ).fetchall()

    for row in rows:
        try:
            conn.execute(
                """INSERT INTO matches (img_id, cad_id, status, method, confidence)
                   VALUES (?, ?, 'auto', 'same-folder', '0.7')""",
                (row["img_id"], row["cad_id"])
            )
            matches_created += 1
        except Exception:
            pass  # Skip duplicates

    # ── Same base filename: images ↔ CAD ──
    rows = conn.execute(
        """SELECT i.img_id, i.filename, c.cad_id, c.filename as cad_filename
           FROM images i
           JOIN cad_files c ON (
             LOWER(SUBSTR(i.filename, 1, INSTR(i.filename || '.', '.') - 1))
             = LOWER(SUBSTR(c.filename, 1, INSTR(c.filename || '.', '.') - 1))
           )
           WHERE i.source_type = 'file_image'
             AND i.img_id NOT IN (SELECT img_id FROM matches WHERE cad_id = c.cad_id)
             AND i.folder != c.folder
           LIMIT 500"""
    ).fetchall()

    for row in rows:
        base_i = os.path.splitext(row["filename"] or "")[0].lower()
        base_c = os.path.splitext(row["cad_filename"] or "")[0].lower()
        if base_i == base_c and len(base_i) >= 3:
            try:
                conn.execute(
                    """INSERT INTO matches (img_id, cad_id, status, method, confidence)
                       VALUES (?, ?, 'auto', 'filename-match', '0.85')""",
                    (row["img_id"], row["cad_id"])
                )
                matches_created += 1
            except Exception:
                pass

    # ── Excel embedded images → Excel records ──
    rows = conn.execute(
        """SELECT i.img_id, i.file_path as excel_path
           FROM images i
           WHERE i.source_type = 'excel_embedded'
             AND i.ex_ref IS NULL OR i.ex_ref = ''"""
    ).fetchall()

    for row in rows:
        try:
            ex_rows = conn.execute(
                "SELECT ex_id FROM excel_records WHERE file_path = ? LIMIT 1",
                (row["excel_path"],)
            ).fetchall()
            if ex_rows:
                ex_id = ex_rows[0]["ex_id"]
                conn.execute(
                    "UPDATE images SET ex_ref = ? WHERE img_id = ?",
                    (ex_id, row["img_id"])
                )
                conn.execute(
                    """INSERT OR IGNORE INTO matches (img_id, ex_id, status, method, confidence)
                       VALUES (?, ?, 'auto', 'excel-reference', '0.9')""",
                    (row["img_id"], ex_id)
                )
                matches_created += 1
        except Exception:
            pass

    # ── PDFs ↔ similar-named images ──
    rows = conn.execute(
        """SELECT p.doc_id, p.filename as pdf_filename, p.folder,
                  i.img_id, i.filename as img_filename
           FROM pdf_files p
           JOIN images i ON p.folder = i.folder
           WHERE i.img_id NOT IN (SELECT img_id FROM matches WHERE pdf_id = p.doc_id)
           LIMIT 500"""
    ).fetchall()

    for row in rows:
        base_p = os.path.splitext(row["pdf_filename"] or "")[0].lower()
        base_i = os.path.splitext(row["img_filename"] or "")[0].lower()
        similarity = _name_similarity(base_p, base_i)
        if similarity > 0.6:
            try:
                conn.execute(
                    """INSERT INTO matches (img_id, pdf_id, status, method, confidence)
                       VALUES (?, ?, 'auto', 'name-similar', ?)""",
                    (row["img_id"], row["doc_id"], str(round(similarity, 2)))
                )
                matches_created += 1
            except Exception:
                pass

    if matches_created > 0:
        conn.commit()
        sys.stderr.write(f"[scan] auto-associated {matches_created} matches\n")
        sys.stderr.flush()

    return matches_created


def _name_similarity(a: str, b: str) -> float:
    """Simple similarity based on common prefix and token overlap."""
    if not a or not b:
        return 0.0
    # Tokenize on common delimiters
    tokens_a = set(re.split(r'[-_\s.]', a.lower()))
    tokens_b = set(re.split(r'[-_\s.]', b.lower()))
    if not tokens_a or not tokens_b:
        return 0.0
    intersection = tokens_a & tokens_b
    union = tokens_a | tokens_b
    return len(intersection) / len(union)


def _auto_index_images(conn, limit: int = 100):
    """Extract CLIP features for images that don't have vector embeddings yet."""
    rows = conn.execute(
        """SELECT i.img_id, i.file_path FROM images i
           LEFT JOIN vector_embeddings ve ON i.img_id = ve.img_id
           WHERE ve.img_id IS NULL
           LIMIT ?""",
        (limit,)
    ).fetchall()

    if not rows:
        return 0

    sys.stderr.write(f"[scan] auto-indexing {len(rows)} images...\n")
    sys.stderr.flush()

    try:
        from backend.services.ai_service import _load_model, _preprocess_image, _extract_features
        _load_model()
    except Exception as e:
        sys.stderr.write(f"[scan] cannot load AI model for indexing: {e}\n")
        sys.stderr.flush()
        return 0

    indexed = 0
    for row in rows:
        img_id = row["img_id"]
        file_path = row["file_path"]
        try:
            if not os.path.exists(file_path):
                continue
            with open(file_path, "rb") as f:
                image_bytes = f.read()
            tensor = _preprocess_image(image_bytes)
            vector = _extract_features(tensor)

            # Save to vector_embeddings
            import numpy as np
            blob = vector.astype(np.float32).tobytes()
            conn.execute(
                "INSERT OR REPLACE INTO vector_embeddings (img_id, vector_dim, vector_blob) VALUES (?, ?, ?)",
                (img_id, len(vector), blob)
            )
            indexed += 1
        except Exception as e:
            sys.stderr.write(f"[scan] index error {img_id}: {e}\n")
            sys.stderr.flush()
            continue

    if indexed > 0:
        conn.commit()
        sys.stderr.write(f"[scan] auto-indexed {indexed} images\n")
        sys.stderr.flush()

    return indexed


def _load_scan_extensions(conn) -> set:
    """Read scan_extensions from settings DB, return lowercase set. Defaults to Excel+PRT."""
    row = conn.execute("SELECT value FROM settings WHERE key = 'scan_extensions'").fetchone()
    if row and row["value"]:
        try:
            exts = json.loads(row["value"])
            if isinstance(exts, list) and len(exts) > 0:
                return {e.lower() for e in exts if isinstance(e, str)}
        except (json.JSONDecodeError, TypeError):
            pass
    return {'.xlsx', '.xls', '.prt'}


def scan_library(library_id: int, library_path: str):
    global _scan_start_time
    _scan_start_time = time.time()
    start_time = _scan_start_time
    path = Path(library_path)

    if not path.exists():
        _add_activity_log("error", "scan", f"Scan failed: path does not exist: {library_path}")
        emit_result({"error": f"Path does not exist: {library_path}", "added": 0, "removed": 0,
                     "modified": 0, "moved": 0, "errors": 1, "duration_sec": 0,
                     "total_files": 0, "excel_count": 0, "image_count": 0,
                     "cad_count": 0, "pdf_count": 0, "other_count": 0})
        return

    if not path.is_dir():
        _add_activity_log("error", "scan", f"Scan failed: path is not a directory: {library_path}")
        emit_result({"error": f"Path is not a directory: {library_path}", "added": 0, "removed": 0,
                     "modified": 0, "moved": 0, "errors": 1, "duration_sec": 0,
                     "total_files": 0, "excel_count": 0, "image_count": 0,
                     "cad_count": 0, "pdf_count": 0, "other_count": 0})
        return

    # Read scan extensions config
    conn = get_connection()
    scan_exts = _load_scan_extensions(conn)

    # Log scan start
    _add_activity_log("info", "scan", f"Scan started: library {library_id} ({library_path})")

    # Phase 1: Collect all file paths (filtered by configured extensions)
    emit_progress("collecting", 0, 0, str(path))
    all_files = []
    try:
        for root, dirs, files in os.walk(path):
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for f in files:
                if not f.startswith('.') and not f.startswith('~'):
                    ext = os.path.splitext(f)[1].lower()
                    if ext in scan_exts:
                        all_files.append(os.path.join(root, f))
    except PermissionError as e:
        _add_activity_log("error", "scan", f"Scan failed: permission denied: {e}")
        emit_result({"error": f"Permission denied: {e}", "added": 0, "removed": 0,
                     "modified": 0, "moved": 0, "errors": 1, "duration_sec": 0,
                     "total_files": 0, "excel_count": 0, "image_count": 0,
                     "cad_count": 0, "pdf_count": 0, "other_count": 0})
        return

    total_files = len(all_files)
    if total_files == 0:
        _add_activity_log("warn", "scan", f"Scan aborted: no matching files found in {library_path}")
        emit_result({"added": 0, "removed": 0, "modified": 0, "moved": 0, "errors": 0,
                     "duration_sec": round(time.time() - start_time, 2),
                     "total_files": 0, "excel_count": 0, "image_count": 0,
                     "cad_count": 0, "pdf_count": 0, "other_count": 0})
        return

    _add_activity_log("info", "scan", f"Collected {total_files} files for scanning")
    # Phase 2: Hash and classify each file
    current_scan = {}
    stats = {"excel": 0, "image": 0, "cad": 0, "pdf": 0, "other": 0}
    errors = 0

    for i, filepath in enumerate(all_files):
        emit_progress("hashing", i + 1, total_files, filepath)
        ext = os.path.splitext(filepath)[1]
        file_type = classify_file(ext)

        try:
            file_size = os.path.getsize(filepath)
            file_hash = sha256_file(filepath)
            stats[file_type] += 1
            current_scan[filepath] = {
                "hash": file_hash,
                "type": file_type,
                "size": file_size,
                "ext": ext,
            }
        except (PermissionError, OSError) as e:
            errors += 1
            sys.stderr.write(f"[scan] error reading {filepath}: {e}\n")
            sys.stderr.flush()

    if errors > 0:
        _add_activity_log("warn", "scan", f"Hashing complete: {len(current_scan)} files hashed with {errors} errors")
    # Phase 3: Change detection — query all indexed tables
    emit_progress("comparing", 0, 0, "")
    conn = get_connection()

    prev_map = {}
    for table in ["images", "cad_files", "pdf_files"]:
        prev_rows = conn.execute(
            f"SELECT file_path, file_hash FROM {table} WHERE file_path LIKE ?",
            (library_path + "%",)
        ).fetchall()
        for row in prev_rows:
            prev_map[row["file_path"]] = row["file_hash"]

    prev_paths = set(prev_map.keys())
    curr_paths = set(current_scan.keys())

    added = curr_paths - prev_paths
    removed = prev_paths - curr_paths
    common = curr_paths & prev_paths

    modified = set()
    for p in common:
        if current_scan[p]["hash"] != prev_map[p]:
            modified.add(p)

    # Detect moved/renamed (same hash, different path)
    curr_hash_to_paths = {}
    for p in added:
        h = current_scan[p]["hash"]
        curr_hash_to_paths.setdefault(h, []).append(p)

    prev_hash_to_paths = {}
    for p in removed:
        h = prev_map[p]
        prev_hash_to_paths.setdefault(h, []).append(p)

    moved_pairs = []
    for h, new_paths in list(curr_hash_to_paths.items()):
        if h in prev_hash_to_paths:
            old_paths = prev_hash_to_paths[h]
            for old_p, new_p in zip(old_paths, new_paths):
                moved_pairs.append((old_p, new_p))
                added.discard(new_p)
                removed.discard(old_p)

    move_count = len(moved_pairs)

    # Log change detection summary
    change_summary_parts = []
    if added: change_summary_parts.append(f"{len(added)} added")
    if removed: change_summary_parts.append(f"{len(removed)} removed")
    if modified: change_summary_parts.append(f"{len(modified)} modified")
    if move_count: change_summary_parts.append(f"{move_count} moved")
    if change_summary_parts:
        _add_activity_log("info", "scan", f"Changes detected: {', '.join(change_summary_parts)}")
    else:
        _add_activity_log("info", "scan", "No changes detected")

    # Phase 4: Save to database
    total_to_save = len(added) + len(removed)
    emit_progress("saving", 0, max(total_to_save, 1), "")

    saved = 0
    for i, filepath in enumerate(added):
        info = current_scan[filepath]
        if info["type"] == "image":
            _insert_image_record(conn, filepath, info)
        elif info["type"] == "cad":
            _insert_cad_record(conn, filepath, info)
        elif info["type"] == "pdf":
            _insert_pdf_record(conn, filepath, info)
        saved += 1
        emit_progress("saving", saved, total_to_save, filepath)

    # Remove deleted files from all tables
    for filepath in removed:
        for table in ["images", "cad_files", "pdf_files"]:
            conn.execute(f"DELETE FROM {table} WHERE file_path = ?", (filepath,))
        saved += 1
        emit_progress("saving", saved, total_to_save, filepath)

    # Log changes
    for p in added:
        _log_change(conn, "added", p)
    for p in removed:
        _log_change(conn, "removed", p)
    for p in modified:
        _log_change(conn, "modified", p, old_value=prev_map[p], new_value=current_scan[p]["hash"])
    for old_p, new_p in moved_pairs:
        _log_change(conn, "moved", new_p, old_value=old_p, new_value=new_p)

    # Extract Excel images
    excel_image_count = 0
    for filepath in curr_paths:
        info = current_scan[filepath]
        if info["type"] == "excel":
            try:
                imgs = _extract_excel_images(filepath, library_path)
                excel_image_count += imgs
            except Exception as e:
                sys.stderr.write(f"[scan] excel error {filepath}: {e}\n")
                sys.stderr.flush()

    # Phase 5: UG preview extraction
    emit_progress("ug_preview", 0, 0, "")
    ug_extracted = 0
    try:
        from backend.services.ug_service import process_directory as ug_process
        ug_result = ug_process(library_path)
        ug_extracted = ug_result.get("extracted", 0)
        if ug_extracted > 0:
            _add_activity_log("info", "scan", f"UG previews extracted: {ug_extracted}")
    except Exception as e:
        _add_activity_log("warn", "scan", f"UG preview extraction failed: {e}")
        sys.stderr.write(f"[scan] UG preview error: {e}\n")
        sys.stderr.flush()

    # Phase 6: Auto-associate images with CAD/Excel/PDF
    emit_progress("matching", 0, 0, "")
    try:
        auto_matches = _auto_associate(conn, library_path)
        if auto_matches > 0:
            _add_activity_log("info", "scan", f"Auto-associated {auto_matches} matches")
    except Exception as e:
        _add_activity_log("warn", "scan", f"Auto-association failed: {e}")
        sys.stderr.write(f"[scan] auto-associate error: {e}\n")
        sys.stderr.flush()
        auto_matches = 0

    # Phase 7: Auto-index new images (extract CLIP features → FAISS)
    emit_progress("indexing", 0, 0, "")
    try:
        auto_indexed = _auto_index_images(conn, limit=200)
        if auto_indexed > 0:
            _add_activity_log("info", "scan", f"Auto-indexed {auto_indexed} images")
    except Exception as e:
        _add_activity_log("warn", "scan", f"Auto-indexing failed: {e}")
        sys.stderr.write(f"[scan] auto-index error: {e}\n")
        sys.stderr.flush()
        auto_indexed = 0

    # Write scan history
    duration = time.time() - start_time
    conn.execute(
        """INSERT INTO scan_history
           (library_id, scan_type, added, removed, modified, moved, errors, duration_sec)
           VALUES (?, 'full', ?, ?, ?, ?, ?, ?)""",
        (library_id, len(added), len(removed), len(modified), move_count, errors, int(duration))
    )

    # Update library record
    conn.execute(
        """UPDATE libraries SET
           file_count = ?, image_count = ?, last_scan = datetime('now','localtime'), status = 'ready'
           WHERE id = ?""",
        (total_files, stats["image"] + excel_image_count, library_id)
    )
    conn.commit()

    # Log scan completion
    scan_errors = errors + (1 if excel_image_count == 0 and stats["excel"] > 0 else 0)
    if scan_errors > 0:
        _add_activity_log("warn", "scan", f"Scan complete with {scan_errors} error(s): {len(added)} added, {len(removed)} removed, {len(modified)} modified, {move_count} moved in {round(duration, 1)}s")
    else:
        _add_activity_log("info", "scan", f"Scan complete: {len(added)} added, {len(removed)} removed, {len(modified)} modified, {move_count} moved in {round(duration, 1)}s")

    emit_result({
        "added": len(added),
        "removed": len(removed),
        "modified": len(modified),
        "moved": move_count,
        "errors": errors,
        "duration_sec": round(duration, 2),
        "total_files": total_files,
        "excel_count": stats["excel"],
        "image_count": stats["image"],
        "cad_count": stats["cad"],
        "pdf_count": stats["pdf"],
        "other_count": stats["other"],
        "excel_image_count": excel_image_count,
        "auto_matches": auto_matches,
        "auto_indexed": auto_indexed,
        "ug_extracted": ug_extracted,
    })


def check_changes(library_id: int) -> dict:
    """Quick check for new/modified/removed files without running a full scan.

    Walks the library directory, hashes files, and compares against the DB.
    Returns counts only — no DB writes, no image extraction, no indexing.
    """
    conn = get_connection()

    lib = conn.execute(
        "SELECT path FROM libraries WHERE id = ?", (library_id,)
    ).fetchone()
    if not lib:
        return {"error": f"Library not found: {library_id}", "added": 0, "removed": 0,
                "modified": 0, "moved": 0, "has_changes": False}

    library_path = lib["path"]
    path = Path(library_path)

    if not path.exists() or not path.is_dir():
        return {"error": f"Path does not exist: {library_path}", "added": 0, "removed": 0,
                "modified": 0, "moved": 0, "has_changes": False}

    scan_exts = _load_scan_extensions(conn)

    # Collect files
    all_files = []
    try:
        for root, dirs, files in os.walk(path):
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for f in files:
                if not f.startswith('.') and not f.startswith('~'):
                    ext = os.path.splitext(f)[1].lower()
                    if ext in scan_exts:
                        all_files.append(os.path.join(root, f))
    except PermissionError:
        return {"error": "Permission denied reading library path", "added": 0, "removed": 0,
                "modified": 0, "moved": 0, "has_changes": False}

    if not all_files:
        # No files on disk — check if DB has old records (all removed)
        prev_count = 0
        for table in ["images", "cad_files", "pdf_files"]:
            c = conn.execute(
                f"SELECT COUNT(*) as n FROM {table} WHERE file_path LIKE ?",
                (library_path + "%",)
            ).fetchone()
            prev_count += c["n"]
        return {"added": 0, "removed": prev_count, "modified": 0, "moved": 0,
                "has_changes": prev_count > 0}

    # Hash each file
    current_scan = {}
    for filepath in all_files:
        try:
            file_hash = sha256_file(filepath)
            current_scan[filepath] = file_hash
        except (PermissionError, OSError):
            pass

    # Query existing records
    prev_map = {}
    for table in ["images", "cad_files", "pdf_files"]:
        prev_rows = conn.execute(
            f"SELECT file_path, file_hash FROM {table} WHERE file_path LIKE ?",
            (library_path + "%",)
        ).fetchall()
        for row in prev_rows:
            prev_map[row["file_path"]] = row["file_hash"]

    prev_paths = set(prev_map.keys())
    curr_paths = set(current_scan.keys())

    added = curr_paths - prev_paths
    removed = prev_paths - curr_paths
    common = curr_paths & prev_paths

    modified = set()
    for p in common:
        if current_scan[p] != prev_map[p]:
            modified.add(p)

    # Detect moved/renamed files
    curr_hash_to_paths = {}
    for p in added:
        h = current_scan[p]
        curr_hash_to_paths.setdefault(h, []).append(p)

    prev_hash_to_paths = {}
    for p in removed:
        h = prev_map[p]
        prev_hash_to_paths.setdefault(h, []).append(p)

    moved_pairs = []
    for h, new_paths in list(curr_hash_to_paths.items()):
        if h in prev_hash_to_paths:
            old_paths = prev_hash_to_paths[h]
            for old_p, new_p in zip(old_paths, new_paths):
                moved_pairs.append((old_p, new_p))
                added.discard(new_p)
                removed.discard(old_p)

    total_changes = len(added) + len(removed) + len(modified) + len(moved_pairs)

    return {
        "added": len(added),
        "removed": len(removed),
        "modified": len(modified),
        "moved": len(moved_pairs),
        "has_changes": total_changes > 0,
        "total_files": len(current_scan),
    }


def execute(method: str, params: dict):
    """JSON-RPC service entry point for scan_service."""
    if method == "scan.checkChanges":
        library_id = params.get("library_id", 0)
        if not isinstance(library_id, int) or library_id <= 0:
            raise ValueError("library_id must be a positive integer")
        return check_changes(library_id)
    else:
        raise ValueError(f"Unknown scan method: {method}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="ZOOBET file scan engine")
    parser.add_argument("--library-id", type=int, required=True)
    parser.add_argument("--path", required=True)
    args = parser.parse_args()
    scan_library(args.library_id, args.path)
