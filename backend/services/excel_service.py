"""Excel parsing service - extract data and embedded images.

Provides functions to:
- Parse Excel files (xlsx/xls/xlsm)
- Extract all worksheet data
- Detect and extract embedded images
- Generate IMG-XXXXXX IDs
- Write records to images and excel_records tables
"""

import sys
import os
import time
import json

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from backend.db.connection import get_connection


def execute(method: str, params: dict):
    if method == "excel.parse":
        return _parse(params.get("file_path", ""), params.get("library_path", ""))
    elif method == "excel.listRecords":
        return _list_records(params.get("file_path"))
    elif method == "excel.extractImages":
        return _extract_images(params.get("file_path", ""), params.get("library_path", ""))
    else:
        raise ValueError(f"Unknown excel method: {method}")


def _parse(file_path: str, library_path: str = "") -> dict[str, object]:
    """Parse an Excel file and write records to excel_records table."""
    if not file_path or not os.path.exists(file_path):
        raise ValueError(f"Excel file not found: {file_path}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

    conn = get_connection()
    folder = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(file_path)))
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    import hashlib
    h = hashlib.sha256()
    with open(file_path, 'rb') as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h.update(chunk)
    file_hash = h.hexdigest()

    wb = load_workbook(file_path, data_only=True)
    total_rows = 0
    sheet_count = 0
    image_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_count += 1
        has_images = 1 if getattr(ws, '_images', []) else 0

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    total_rows += 1
                    col_letter = _col_letter(col_idx)
                    ex_id = f"EX-{os.path.basename(file_path)}-{sheet_name}-R{row_idx}C{col_idx}"[:128]

                    cell_str = str(cell.value)[:65535] if cell.value is not None else None

                    conn.execute(
                        """INSERT OR REPLACE INTO excel_records
                           (ex_id, file_path, folder, filename, sheet_name, row_number, column_name, cell_value, has_image, file_hash, last_modified, indexed_at)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (ex_id, file_path, folder, filename, sheet_name, row_idx, col_letter,
                         cell_str, has_images, file_hash, mtime, now)
                    )

    wb.close()

    conn.commit()

    return {
        "file_path": file_path,
        "sheets": sheet_count,
        "rows": total_rows,
        "images": image_count,
        "file_hash": file_hash,
    }


def _list_records(file_path: str | None = None) -> list:
    conn = get_connection()
    if file_path:
        rows = conn.execute(
            "SELECT * FROM excel_records WHERE file_path = ? ORDER BY sheet_name, row_number LIMIT 500",
            (file_path,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT DISTINCT file_path, filename, sheet_name, COUNT(*) as cnt, MAX(indexed_at) as last_idx "
            "FROM excel_records GROUP BY file_path ORDER BY last_idx DESC LIMIT 100"
        ).fetchall()
    return [dict(row) for row in rows]


def _extract_images(file_path: str, library_path: str = "") -> dict:
    """Extract embedded images from an Excel file."""
    if not file_path or not os.path.exists(file_path):
        raise ValueError(f"Excel file not found: {file_path}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    extracted_dir = os.path.join(_PROJECT_ROOT, "backend", "data", "extracted_images")
    os.makedirs(extracted_dir, exist_ok=True)

    conn = get_connection()
    folder = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(file_path)))
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    wb = load_workbook(file_path, data_only=True)
    extracted = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        images = getattr(ws, '_images', [])
        if not images:
            continue

        for img in images:
            try:
                # openpyxl >= 3.1 dropped .image; use _data() as fallback
                if hasattr(img, 'image'):
                    pil_img = img.image
                else:
                    from io import BytesIO
                    from PIL import Image as _PILImage
                    pil_img = _PILImage.open(BytesIO(img._data()))
                if pil_img is None:
                    continue

                fmt = pil_img.format if pil_img.format else "PNG"

                # Generate sequential IMG ID
                max_row = conn.execute(
                    "SELECT img_id FROM images WHERE img_id LIKE 'IMG-%' ORDER BY img_id DESC LIMIT 1"
                ).fetchone()
                if max_row:
                    last_num = int(max_row["img_id"].split("-")[1])
                    img_id = f"IMG-{last_num + 1:06d}"
                else:
                    img_id = "IMG-000001"

                ext = fmt.lower()
                if ext == "jpeg":
                    ext = "jpg"

                img_filename = f"{img_id}.{ext}"
                img_path = os.path.join(extracted_dir, img_filename)
                pil_img.save(img_path, format=fmt)

                img_size = os.path.getsize(img_path)
                # Extract cell ref from anchor (avoid str() — broken in openpyxl 3.1+Python 3.13)
                cell_ref = ''
                if hasattr(img, 'anchor') and img.anchor is not None:
                    try:
                        a = img.anchor
                        if hasattr(a, '_from'):
                            col_letter = _col_letter(a._from.col + 1)
                            cell_ref = f"{col_letter}{a._from.row + 1}"
                    except Exception:
                        pass

                conn.execute(
                    """INSERT OR REPLACE INTO images
                       (img_id, source_type, file_path, folder, filename, size_bytes, width, height, file_hash, ex_ref, status, last_modified, indexed_at)
                       VALUES (?, 'excel_embedded', ?, ?, ?, ?, ?, ?, NULL, ?, 'normal', ?, ?)""",
                    (img_id, file_path, folder, img_filename, img_size, pil_img.width, pil_img.height, cell_ref, mtime, now)
                )

                extracted.append({
                    "img_id": img_id,
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "filename": img_filename,
                    "width": pil_img.width,
                    "height": pil_img.height,
                    "size_bytes": img_size,
                })

            except Exception as e:
                sys.stderr.write(f"[excel] extract error {file_path} sheet={sheet_name}: {e}\n")
                sys.stderr.flush()
                continue

    wb.close()
    conn.commit()

    return {
        "file_path": file_path,
        "extracted": len(extracted),
        "images": extracted,
    }


def _col_letter(n: int) -> str:
    """Convert column number to Excel column letter(s). 1->A, 27->AA."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result
