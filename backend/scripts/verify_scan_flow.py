"""扫描流程端到端验证脚本

验证: db.init → library.add → Excel parse → image extract → images table

Run: python backend/scripts/verify_scan_flow.py [--test-data ./test-data]
Uses a temporary database, does NOT modify the production database.
Exit code: 0 = all checks passed, non-0 = failure.
Outputs JSON result on the last line.
"""

import os
import sys
import json
import tempfile
import shutil
import argparse
import time
import traceback

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# ── Database isolation ────────────────────────────────────────────────
# Patch connection module BEFORE any service imports use get_connection()

_TMP_DB_DIR = tempfile.mkdtemp(prefix="zoobet_verify_scan_")
_TMP_DB_PATH = os.path.join(_TMP_DB_DIR, "zoobet.db")

import backend.db.connection as _conn

_conn._DB_DIR = _TMP_DB_DIR
_conn._DB_PATH = _TMP_DB_PATH
os.makedirs(_TMP_DB_DIR, exist_ok=True)

# Close any cached production connection, so next get_connection() uses our temp db
_conn.close_connection()

# Now safe to import services — their get_connection() will use the temp db
from backend.services import excel_service, library_service  # noqa: E402
from backend.db.connection import init_schema, get_connection  # noqa: E402


# ── Helpers ───────────────────────────────────────────────────────────

results = []
failures = 0


def check(name: str, ok: bool, detail: str = ""):
    global failures
    status = "PASS" if ok else "FAIL"
    msg = f"  [{status}] {name}"
    if detail:
        msg += f" -- {detail}"
    print(msg)
    results.append({"name": name, "ok": ok, "detail": detail})
    if not ok:
        failures += 1


def create_excel_with_image(path: str) -> bool:
    """Create an Excel file with an embedded test image.

    Returns True if successful, False if openpyxl is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from PIL import Image as PilImage
    except ImportError:
        return False

    os.makedirs(os.path.dirname(path), exist_ok=True)

    # Create a small test PNG in memory
    from io import BytesIO
    pil_img = PilImage.new("RGB", (64, 64), color=(255, 128, 0))
    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["A2"] = "001"
    ws["B2"] = "Test Part"

    xl_img = XlImage(buf)
    xl_img.anchor = "C2"
    ws.add_image(xl_img)

    wb.save(path)
    return True


def table_exists(conn, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


# ── Main ──────────────────────────────────────────────────────────────

def main():
    global failures

    parser = argparse.ArgumentParser(description="Scan flow end-to-end verification")
    parser.add_argument("--test-data", default=os.path.join(_PROJECT_ROOT, "test-data"),
                        help="Path to test-data directory")
    parser.add_argument("--keep-db", action="store_true",
                        help="Keep the temporary test database after verification")
    args = parser.parse_args()

    test_data_dir = os.path.abspath(args.test_data)

    print("=" * 60)
    print("ZOOBET Scan Flow Verification")
    print("=" * 60)
    print(f"Project root: {_PROJECT_ROOT}")
    print(f"Test DB:      {_TMP_DB_PATH}")
    print(f"Test data:    {test_data_dir}")
    print()

    try:
        # ── 1. db.init ──────────────────────────────────────────────────
        print("-- 1. db.init --")

        try:
            init_schema()
            check("db.init completes without exception", True)
        except Exception as e:
            check("db.init completes without exception", False, str(e))
            print(f"\n{failures} failure(s). Aborting.")
            sys.exit(1)

        conn = get_connection()

        expected_tables = [
            "cad_files", "change_logs", "excel_records", "images", "libraries",
            "matches", "pdf_files", "scan_history", "search_history", "settings", "vector_embeddings"
        ]
        tables_row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
        actual_tables = [r["name"] for r in tables_row]
        all_present = all(t in actual_tables for t in expected_tables)
        check("All 11 schema tables created", all_present,
              f"Found {len(actual_tables)} tables")
        missing = [t for t in expected_tables if t not in actual_tables]
        if missing:
            check("No missing tables", False, f"Missing: {missing}")

        # ── 2. library.add ──────────────────────────────────────────────
        print("\n-- 2. library.add --")

        try:
            result = library_service.execute("library.add", {
                "path": test_data_dir,
                "label": "Verify Scan Flow Test"
            })
            lib_id = result.get("id")
            check("library.add returns id", lib_id is not None, str(result))
            check("library.add path matches", result.get("path") == test_data_dir,
                  f"got: {result.get('path')}")
        except Exception as e:
            check("library.add succeeds", False, str(e))

        # Verify library in database
        rows = conn.execute("SELECT * FROM libraries WHERE path = ?", (test_data_dir,)).fetchall()
        check("Library record exists in DB", len(rows) == 1,
              f"Found {len(rows)} rows")
        if rows:
            lib = dict(rows[0])
            check("Library status is idle", lib.get("status") == "idle",
                  f"status={lib.get('status')}")

        # ── 3. Excel parsing ────────────────────────────────────────────
        print("\n-- 3. Excel parsing --")

        test_excel = os.path.join(test_data_dir, "excel", "parts_catalog.xlsx")
        if not os.path.exists(test_excel):
            print(f"  Test Excel not found: {test_excel}")
            print(f"  Run seed script first: python backend/scripts/seed_test_data.py --target {test_data_dir}")
            check("Test Excel file exists", False, test_excel)
        else:
            check("Test Excel file exists", True, test_excel)

            try:
                parse_result = excel_service.execute("excel.parse", {
                    "file_path": test_excel,
                    "library_path": test_data_dir,
                })
                check("excel.parse completes without exception", True,
                      f"sheets={parse_result.get('sheets')}, rows={parse_result.get('rows')}")
                check("excel.parse returns file_path", parse_result.get("file_path") == test_excel)
                check("excel.parse returns sheets > 0", parse_result.get("sheets", 0) > 0,
                      f"sheets={parse_result.get('sheets')}")
                check("excel.parse returns rows > 0", parse_result.get("rows", 0) > 0,
                      f"rows={parse_result.get('rows')}")
            except Exception as e:
                check("excel.parse completes without exception", False, str(e))

            # Verify excel_records table has data
            rec_count = conn.execute(
                "SELECT COUNT(*) as n FROM excel_records WHERE file_path = ?",
                (test_excel,)
            ).fetchone()["n"]
            check("excel_records has rows after parse", rec_count > 0,
                  f"{rec_count} records in excel_records")

        # ── 4. Image extraction ─────────────────────────────────────────
        print("\n-- 4. Image extraction --")

        # The seed test Excel has NO embedded images, so we create one that does
        test_excel_with_image = os.path.join(_TMP_DB_DIR, "test_with_image.xlsx")
        created = create_excel_with_image(test_excel_with_image)
        if not created:
            check("Create Excel with embedded image", False,
                  "openpyxl or Pillow not available — cannot test image extraction")
        else:
            check("Create Excel with embedded image", True, test_excel_with_image)

            try:
                extract_result = excel_service.execute("excel.extractImages", {
                    "file_path": test_excel_with_image,
                    "library_path": test_data_dir,
                })
                extracted_count = extract_result.get("extracted", 0)
                check("excel.extractImages completes without exception", True,
                      f"extracted={extracted_count}")
                check("excel.extractImages found images", extracted_count > 0,
                      f"extracted {extracted_count} images")
            except Exception as e:
                check("excel.extractImages completes without exception", False, str(e))

            # Verify images table has records from extraction
            img_rows = conn.execute(
                "SELECT * FROM images WHERE source_type = 'excel_embedded'"
            ).fetchall()
            img_count = len(img_rows)
            check("images table has excel_embedded records", img_count > 0,
                  f"{img_count} rows with source_type=excel_embedded")

            if img_rows:
                img = dict(img_rows[0])
                check("Image has img_id", bool(img.get("img_id")),
                      f"img_id={img.get('img_id')}")
                check("Image has source_type", img.get("source_type") == "excel_embedded")
                check("Image has file_path", bool(img.get("file_path")))
                check("Image has filename", bool(img.get("filename")))
                check("Image has size_bytes > 0", (img.get("size_bytes") or 0) > 0,
                      f"size_bytes={img.get('size_bytes')}")
                check("Image has width > 0", (img.get("width") or 0) > 0,
                      f"width={img.get('width')}")
                check("Image has height > 0", (img.get("height") or 0) > 0,
                      f"height={img.get('height')}")
                check("Image status is normal", img.get("status") == "normal",
                      f"status={img.get('status')}")

            # Clean up extracted images (excel_service writes to real extracted_images dir)
            if created:
                try:
                    for img_info in extract_result.get("images", []):
                        fname = img_info.get("filename", "")
                        if fname:
                            img_file = os.path.join(_PROJECT_ROOT, "backend", "data", "extracted_images", fname)
                            if os.path.exists(img_file):
                                os.remove(img_file)
                except Exception:
                    pass

        # ── 5. Index status queryable ────────────────────────────────────
        print("\n-- 5. Index status --")

        # Verify we can query image counts by source type
        source_counts = conn.execute(
            "SELECT source_type, COUNT(*) as n FROM images GROUP BY source_type"
        ).fetchall()
        source_map = {r["source_type"]: r["n"] for r in source_counts}
        check("Image counts by source_type queryable", len(source_map) > 0,
              str(source_map))

        # Verify we can query total images
        total_images = conn.execute("SELECT COUNT(*) as n FROM images").fetchone()["n"]
        check("Total images count queryable", total_images >= 0,
              f"total_images={total_images}")

        # Verify excel_records are queryable
        total_records = conn.execute("SELECT COUNT(*) as n FROM excel_records").fetchone()["n"]
        check("Excel records count queryable", total_records >= 0,
              f"total_records={total_records}")

        # Verify library stats are queryable
        lib_stats = conn.execute(
            "SELECT COUNT(*) as n, SUM(file_count) as files, SUM(image_count) as imgs FROM libraries"
        ).fetchone()
        check("Library aggregate stats queryable", lib_stats is not None,
              f"libraries={lib_stats['n']}")

        # Verify scan_history is queryable (even if empty)
        sh_count = conn.execute("SELECT COUNT(*) as n FROM scan_history").fetchone()["n"]
        check("scan_history table queryable", sh_count >= 0,
              f"scan_history rows={sh_count}")

    except Exception as e:
        check("Unexpected error", False, f"{e}\n{traceback.format_exc()}")

    finally:
        # ── Cleanup temp database ──────────────────────────────────────
        conn = getattr(_conn._local, "connection", None)
        if conn:
            try:
                conn.close()
                _conn._local.connection = None
            except Exception:
                pass

        if not args.keep_db:
            shutil.rmtree(_TMP_DB_DIR, ignore_errors=True)
            print(f"\n  Temp DB cleaned up: {_TMP_DB_DIR}")
        else:
            print(f"\n  Temp DB kept at: {_TMP_DB_PATH}")

    # ── Summary ────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    total = len(results)
    passed = total - failures
    print(f"Results: {passed}/{total} passed, {failures} failed")
    for r in results:
        status = "PASS" if r["ok"] else "FAIL"
        print(f"  [{status}] {r['name']}")

    summary = {
        "script": "verify_scan_flow.py",
        "total": total,
        "passed": passed,
        "failed": failures,
        "success": failures == 0,
        "results": results,
    }
    # Output machine-readable JSON as the last line on stdout
    print("\n" + json.dumps(summary, ensure_ascii=False))

    print("=" * 60)

    if failures > 0:
        print(f"\n{failures} VERIFICATION FAILURE(S)")
        sys.exit(1)
    else:
        print("\nAll scan flow verifications passed.")
        sys.exit(0)


if __name__ == "__main__":
    main()
