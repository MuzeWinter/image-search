"""Create test data for ZOOBET backend verification.

Generates a folder structure with placeholder files of all supported types.
Run: python backend/scripts/seed_test_data.py [--target ./test-data]
"""

import os
import sys
import argparse
import struct
import zlib

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)


def create_png(path: str, r: int = 255, g: int = 0, b: int = 0):
    """Create a minimal 1x1 PNG with the given RGB color."""
    raw = b"\x00" + struct.pack("BBB", r, g, b)

    def chunk(ctype, data):
        c = ctype + data
        crc = struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
        return struct.pack(">I", len(data)) + c + crc

    ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(b"\x89PNG\r\n\x1a\n")
        f.write(chunk(b"IHDR", ihdr))
        f.write(chunk(b"IDAT", zlib.compress(raw)))
        f.write(chunk(b"IEND", b""))


def create_pdf(path: str, content: str = "Test PDF content"):
    """Create a minimal valid PDF."""
    pdf = (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/MediaBox[0 0 612 792]/Parent 2 0 R>>endobj\n"
        b"xref\n0 4\n0000000000 65535 f \n0000000009 00000 n \n"
        b"0000000058 00000 n \n0000000115 00000 n \n"
        b"trailer<</Size 4/Root 1 0 R>>\nstartxref\n190\n%%EOF"
    )
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(pdf)


def create_excel(path: str):
    """Create a sample Excel workbook with multiple sheets."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("[seed] openpyxl not installed, skipping Excel creation")
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"
    ws["A1"] = "Part Number"
    ws["B1"] = "Description"
    ws["C1"] = "Price"
    ws["A2"] = "PN-001"
    ws["B2"] = "Flange Bracket"
    ws["C2"] = 42.50
    ws["A3"] = "PN-002"
    ws["B3"] = "Mounting Plate"
    ws["C3"] = 18.75

    ws2 = wb.create_sheet("Materials")
    ws2["A1"] = "Material"
    ws2["A2"] = "Steel"
    ws2["A3"] = "Aluminum"

    wb.save(path)
    print(f"[seed] Created Excel: {path}")


def main():
    parser = argparse.ArgumentParser(description="ZOOBET test data generator")
    parser.add_argument("--target", default="./test-data", help="Target directory")
    args = parser.parse_args()

    target = os.path.abspath(args.target)
    print(f"[seed] Creating test data in: {target}")

    # Directory structure
    dirs = ["images", "excel", "cad", "pdf"]
    for d in dirs:
        os.makedirs(os.path.join(target, d), exist_ok=True)

    # Placeholder images (different colors for visual distinction)
    create_png(os.path.join(target, "images", "photo1.png"), 255, 100, 100)
    create_png(os.path.join(target, "images", "photo2.png"), 100, 255, 100)
    create_png(os.path.join(target, "images", "diagram.png"), 100, 100, 255)
    print("[seed] Created 3 placeholder PNG images")

    # CAD files (minimal DXF and STEP)
    os.makedirs(os.path.join(target, "cad"), exist_ok=True)
    dxf_content = (
        "0\nSECTION\n2\nHEADER\n9\n$ACADVER\n1\nAC1021\n0\nENDSEC\n0\nEOF\n# Bracket drawing"
    )
    step_content = (
        "ISO-10303-21;\nHEADER;\nFILE_DESCRIPTION('Assembly Model');\n"
        "ENDSEC;\nDATA;\n#1=CARTESIAN_POINT((0,0,0));\nENDSEC;\nEND-ISO-10303-21;"
    )
    with open(os.path.join(target, "cad", "bracket.dxf"), "w") as f:
        f.write(dxf_content)
    with open(os.path.join(target, "cad", "assembly.step"), "w") as f:
        f.write(step_content)
    print("[seed] Created 2 CAD files")

    # PDF
    create_pdf(os.path.join(target, "pdf", "spec_sheet.pdf"))
    print("[seed] Created 1 PDF file")

    # Excel
    create_excel(os.path.join(target, "excel", "parts_catalog.xlsx"))

    # Other (for classification testing)
    with open(os.path.join(target, "readme.txt"), "w") as f:
        f.write("ZOOBET test data readme\nThis file tests 'other' classification.\n")
    print("[seed] Created 1 text file (other)")

    print(f"\n[seed] Done. Created {9} test files in {target}")
    print("[seed] Run: python backend/services/scan_service.py --library-id 1 --path " + target)


if __name__ == "__main__":
    main()
